import glob
import os
import re
import logging
import polars as pl
import datetime as dt
from typing import AnyStr, List, Tuple, Dict, Union
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,Alignment, Border, Side
from openpyxl.utils import get_column_letter
import concurrent.futures

class FileManager:
    def __init__(self, base_dir: str):
        self.base_dir = base_dir
        self._output_path = None
        # 设置日志配置
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


    @property
    def output_path(self):
        ### 获取输出路径 ###
        return self._output_path

    @property
    def base_directory(self):
        ### 获取基础目录 ###
        return self.base_dir


    def save_to_sheet(self,filename:str,formatter=None,**sheets) -> str:
        """
        将多个 Polars DataFrame 保存到一个 Excel 文件的多个 sheet
        :param filename: 文件名
        :param formatter: 格式化器类（可选）
        :param sheets: 包含多个sheet的数据
        :return: 输出文件路径
        """
        try:
            print("Debug - Sheets content:", sheets)  # 添加调试信息
            self._output_path = self._generate_output_path(filename)

            workbook = Workbook()

            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            if not sheets:
                worksheet = workbook.create_sheet('Sheet')
                workbook.save(self._output_path)
                return self._output_path
            
            for sheet_name, df in sheets.items():
                if isinstance(df, pl.DataFrame):
                    worksheet = workbook.create_sheet(sheet_name)
                    self._save_dataframe_to_sheet(worksheet, df)
                else:
                    if sheet_name != 'formatter':
                        logging.warning(f"{sheet_name} 不是 Polars DataFrame 类型，无法保存")
           
            # 如果提供了格式化器，则应用格式化
            if formatter:
                formatter_instance = formatter(workbook)
                formatter_instance.apply_formatting()

            workbook.save(self._output_path)
            logging.info(f"文件保存成功，路径：{self._output_path}")
            return self._output_path
        
        except Exception as e:
            logging.error(f"数据保存失败: {e}")
            return None


    def _save_dataframe_to_sheet(self, worksheet, df: pl.DataFrame):
        """封装写入 Polars DataFrame 到 Excel sheet 的方法"""
        try:
            # 调试信息
            logging.info(f"正在保存数据，形状: {df.shape}")
        
            # 设置样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
        
            # 1. 写入列名并设置格式
            for col_idx, col_name in enumerate(df.columns, 1):
                # 去除列名中的后缀
                cleaned_col_name = re.sub(r"_\d+$", "", col_name)
                cell = worksheet.cell(row=1, column=col_idx, value=cleaned_col_name)
                # 设置格式
                cell.border = thin_border
                cell.alignment = center_alignment
        
            # 2. 写入数据并设置格式
            for row_idx, row in enumerate(df.rows(), 1):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx+1, column=col_idx)
                    # 处理日期时间格式
                    if isinstance(value, dt.datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    cell.value = value
                    # 设置格式
                    cell.border = thin_border
                    cell.alignment = center_alignment

        except Exception as e:
            logging.error(f"保存数据到工作表时出错: {str(e)}")
            logging.error(f"数据示例: {df.head()}")
            raise
 

    def _generate_output_path(self, filename: str) -> str:
        """生成输出文件的路径，并检查文件是否存在"""
        try:
            date_str = dt.datetime.now().strftime("%Y%m%d")
            output_filename = f"{filename}_预处理_{date_str}.xlsx"
            output_path = os.path.join(self.base_dir, output_filename)

            if os.path.exists(output_path):
                logging.warning(f"文件 {output_path} 已存在，将覆盖")
                os.remove(output_path)

            return output_path
        except Exception as e:
            logging.error(f"生成文件路径失败: {e}")
            return None

    ### 获取basedir下某个文件夹下的文件 ###
    def get_list_files(self, dir_name: str, file_extension="*.xls*") -> list[AnyStr]:
        folder_path = os.path.join(self.base_dir, dir_name)
        if not os.path.exists(folder_path):
            logging.error(f"未找到指定目录: {folder_path}")
            exit(1)
            
        file_list = glob.glob(folder_path + "/" + file_extension)
        return file_list
        
    def get_latest_file(self, dir_name: str, file_extension=(".xlsx", ".xls")) -> str:
        source_dir = os.path.join(self.base_dir, dir_name)
        if not os.path.exists(source_dir):
            logging.error(f"未找到指定目录: {source_dir}")
            exit(1)

        try:
            files = [f for f in os.listdir(source_dir)
                     if os.path.isfile(os.path.join(source_dir, f)) and f.endswith(file_extension)]
            if not files:
                logging.warning(f"目录 {source_dir} 中不存在文件")
                return ""

            latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(source_dir, x)))
            return os.path.join(source_dir, latest_file)
        except Exception as e:
            logging.error(f"获取最新文件时发生错误: {e}")
            return ""

    def read_excel(self,file_name:str = None,file_path: str = None, sheet_name: str = None, show_logs=False) -> pl.DataFrame | dict[str, pl.DataFrame]:
        try:
            if file_path is None and file_name is None:
                logging.error("未提供文件路径或文件名")
                return pl.DataFrame()
            
            if file_path is None and file_name is not None:
                file_path = os.path.join(self.base_dir, file_name)
            elif file_path is not None and file_name is not None:
                file_path = os.path.join(file_path, file_name)

                # 检查文件是否存在
            if not os.path.exists(file_path):
                logging.error(f"文件不存在: {file_path}")
                return pl.DataFrame()
        
             # 修改：移除过度的数据清理
            data = pl.read_excel(file_path, sheet_name=sheet_name)

            if isinstance(data, pl.DataFrame):
                if show_logs:
                    logging.info(f"成功读取 {file_path}")
            elif isinstance(data, dict):
                if show_logs:
                    for sheet_key in data.keys():
                        logging.info(f"成功读取 {file_path} 中的 {sheet_key} 表")
                    
            return data
        except Exception as e:
            logging.error(f"读取文件 {file_path} 中的 sheet: {sheet_name} 失败: {e}")
            return pl.DataFrame() if isinstance(sheet_name, (str, int)) else {}

    def read_csv(self, file_path: str = None, dir_name: str = None, file_name: str = None, separator: str = ",",
                 has_header: bool = True,
                 new_columns: list[str] = None, encoding: str = "utf-8",
                 show_logs: bool = False, **kwargs) -> pl.DataFrame:
        """
        使用 Polars 读取 CSV 文件。

        参数:
        - file_path: CSV 文件的完整路径。
        - dir_name: 目录名，如果提供，则会在 base_dir 下的这个目录中查找 file_name。
        - file_name: 文件名，如果提供，则会在 base_dir 或 dir_name 指定的目录下查找这个文件。
        - separator: 字段分隔符，默认为逗号。
        - has_header: 文件是否包含标题行，默认为 True。
        - new_columns: 可选，为数据列指定新的列名。
        - encoding: 文件编码，默认为 "utf-8"。
        - show_logs: 是否显示读取日志，默认为 False。
        - **kwargs: 传递给 pl.read_csv 的其他关键字参数。

        返回:
        - pl.DataFrame: 读取的数据。
        """
        try:
            if file_path is None:
                if file_name is None:
                    raise ValueError("必须提供 file_path 或 file_name")
                if dir_name:
                    file_path = os.path.join(self.base_dir, dir_name, file_name)
                else:
                    file_path = os.path.join(self.base_dir, file_name)

            if not os.path.exists(file_path):
                logging.error(f"文件不存在: {file_path}")
                return pl.DataFrame()

            data = pl.read_csv(
                file_path,
                separator=separator,
                has_header=has_header,
                new_columns=new_columns,
                encoding=encoding,
                **kwargs
            )

            if show_logs:
                logging.info(f"成功读取 {file_path}")

            return data

        except Exception as e:
            logging.error(f"读取文件 {file_path} 失败: {e}")
            return pl.DataFrame()
            
        
    def save_to_excel(self, df: pl.DataFrame, file_name: str,file_path:str = None):
        try:
            if df is None:
                logging.error("DataFrame为空，无法保存")
                return
            temp_file_name = self._generate_output_path(file_name)

            output_path = None
            
            if file_path is None:
                output_path = temp_file_name
            else:
                output_path = os.path.join(file_path, temp_file_name)
                
            if os.path.exists(output_path):
                logging.warning(f"文件{output_path}已存在，将覆盖")
                os.remove(output_path)
                
            # 创建一个新的工作簿
            workbook = Workbook()
            worksheet = workbook.active

            # 使用封装的方法保存 DataFrame
            self._save_dataframe_to_sheet(worksheet, df)

            # 保存 Excel 文件
            workbook.save(output_path)
            logging.info(f"数据保存成功，路径：{output_path}")
            return output_path
        except Exception as e:
            logging.error(f"数据保存失败: {e}")
            return None
