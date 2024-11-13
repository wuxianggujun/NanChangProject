import polars as pl
from openpyxl import load_workbook
from tqdm import tqdm

def process_datas(file_path):
    df = pl.read_excel(file_path)

    times = {}

    groups = df.group_by('受理号码')
    for row in tqdm(df.to_dicts(), desc='处理数据'):
        phone = row['受理号码']
        date = row['系统接单时间'][:7]
        if phone not in times:
            times[phone] = {date: 1}
        else:
            if date not in times[phone]:
                times[phone][date] = 1
            else:
                times[phone][date] += 1

    datas = df.to_dicts()

    for row in tqdm(datas, desc='插入数据'):
        phone = row['受理号码']
        date = row['系统接单时间'][:7]
        row.update({'当月投诉次数': times[phone][date]})

    result_df = pl.DataFrame(datas)

    data = result_df.to_dicts()

    workbook = load_workbook(file_path)

    worksheet = workbook.create_sheet("统计结果")

    columns = result_df.columns
    for col_num, col_name in enumerate(columns, start=1):
        worksheet.cell(row=1, column=col_num, value=col_name)

    for row_num, row in enumerate(data, start=2):
        for col_num, col_name in enumerate(columns, start=1):
            worksheet.cell(row=row_num, column=col_num, value=row[col_name])

    workbook.save(file_path)

    print(f'数据保存到{file_path}')

if __name__ == '__main__':
    file_path = r"WorkDocument/重复投诉日报/source/工单查询 (60).xlsx"

    process_datas(file_path)