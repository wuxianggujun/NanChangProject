import datetime as dt
import logging
import openpyxl
import polars as pl
from tool.data import DataUtils
from tool.file import FileManager
from tool.ExcelFormatter import ExcelFormatter


def generate_repeat_complaints_table(dataframe: pl.DataFrame) -> pl.DataFrame:
    stats_data = []
    region_order = ["南昌", "九江", "上饶", "抚州", "宜春", "吉安", "赣州", "景德镇", "萍乡", "新余", "鹰潭"]

    # 初始化统计字典，为每个地区创建一个计数字典
    region_stats = {region: {"重复2次": 0, "重复3次": 0, "重复4次及以上": 0} for region in region_order}

    repeat_complaints = dataframe.filter(pl.col("重复投诉次数") >= 2)

    # 对每个区域-受理号码只统计一次
    unique_complaints = repeat_complaints.unique(subset=["区域-受理号码"])

    # 统计每个地区的重复投诉次数
    for row in unique_complaints.iter_rows(named=True):
        region = row["区域"].replace("市", "")
        count = row["重复投诉次数"]
       
        if region in region_stats:
            if count == 2:
                region_stats[region]["重复2次"] += 1
            elif count == 3:
                region_stats[region]["重复3次"] += 1
            elif count >= 4:
                region_stats[region]["重复4次及以上"] += 1
            
    # 转换为DataFrame格式
    for region in region_order:
        # 计算当日新增重复投诉总计
        total = (region_stats[region]["重复2次"] +
                 region_stats[region]["重复3次"] +
                 region_stats[region]["重复4次及以上"])

        # 将0值转换为None（这样在DataFrame中显示为空）
        row_data = {
            "区域": region,
            "重复2次": region_stats[region]["重复2次"] if region_stats[region]["重复2次"] > 0 else None,
            "重复3次": region_stats[region]["重复3次"] if region_stats[region]["重复3次"] > 0 else None,
            "重复4次及以上": region_stats[region]["重复4次及以上"] if region_stats[region][
                                                                          "重复4次及以上"] > 0 else None,
            "当日新增重复投诉总计": total if total > 0 else 0,
            "今天重复投诉解决情况": None,
            "累计重复投诉解决率": None
        }
        
        stats_data.append(row_data)

    # 添加总计行
    totals = {
        "区域": "总计",
        "重复2次": sum(row["重复2次"] or 0 for row in stats_data),
        "重复3次": sum(row["重复3次"] or 0 for row in stats_data),
        "重复4次及以上": sum(row["重复4次及以上"] or 0 for row in stats_data),
        "当日新增重复投诉总计": sum(row["当日新增重复投诉总计"] for row in stats_data),
        "今天重复投诉解决情况": 0,
        "累计重复投诉解决率": None
    }
    stats_data.append(totals)

    stats_df = pl.DataFrame(stats_data)
    return stats_df


def extract_address_and_issue(complaint_content):
    return "", ""


def process_excel(df: pl.DataFrame) -> tuple:
   
       # 打印初始数据信息
    logging.info(f"初始数据形状: {df.shape}")
    logging.info(f"系统接单时间列类型: {df['系统接单时间'].dtype}")
    logging.info(f"系统接单时间样本: {df['系统接单时间'].head()}")

    try:
      
        logging.info("初始数据类型转换完成")

            # 定义时间范围
        start_time = (dt.datetime.now() - dt.timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_time = dt.datetime.now().replace(hour=16, minute=0, second=0, microsecond=0)
      

        logging.info(f"开始处理：{start_time} 到 {end_time} 共三十天的数据...")

        dataframe = DataUtils(df).filter_data_range(
            date_column="系统接单时间",
            start_time=start_time,
            end_time=end_time
        )

        dataframe = DataUtils(dataframe).clean_and_unique(
            unique_columns=["客服流水号"],
            drop_columns=["序号", "工单号"]
        )

        dataframe = DataUtils(dataframe).add_date_only_column(
            date_column="系统接单时间",
            new_column="系统接单时间2",
            format="%Y/%m/%d %H"
        )

        dataframe = DataUtils(dataframe).insert_colum("系统接单时间", "系统接单时间2")

        dataframe = DataUtils(dataframe).drop_columns_after(column_name="口碑未达情况原因")

        dataframe = DataUtils(dataframe).combine_columns(
            columns=["区域", "受理号码"],
            separator="-",
            new_column="区域-受理号码"
        )

        # 添加筛选条件：今天下午四点到昨天下午四点的数据
        yesterday_end = (dt.datetime.now() - dt.timedelta(days=1)).replace(hour=16, minute=0, second=0, microsecond=0)
        today_start = dt.datetime.now().replace(hour=16, minute=0, second=0, microsecond=0)

        # 筛选出时间段内的重复工单
        filtered_df = dataframe.filter(
            (pl.col("系统接单时间") >= yesterday_end) & (pl.col("系统接单时间") <= today_start)
        )

        # 检查是否有数据
        if len(filtered_df) == 0:
            logging.warning(f"在 {yesterday_end} 到 {today_start} 期间没有找到任何投诉数据")
            # 返回空的DataFrame但保持结构一致
            empty_stats = pl.DataFrame({
                "区域": ["总计"],
                "重复2次": [0],
                "重复3次": [0],
                "重复4次及以上": [0],
                "当日新增重复投诉总计": [0],
                "今天重复投诉解决情况": [0],
                "累计重复投诉解决率": [None]
            })
            return dataframe, pl.DataFrame(), pl.DataFrame({"投诉信息": ["无重复投诉数据"]}), empty_stats

        result_list = []

        for row in filtered_df.iter_rows(named=True):
            matches = dataframe.filter(pl.col("区域-受理号码") == row["区域-受理号码"])
            result_list.extend(matches.to_dicts())

        result_df = pl.DataFrame(result_list)

         # 计算重复投诉次数
        repeat_counts = result_df.group_by("区域-受理号码").agg(pl.len().alias("重复投诉次数"))

        result_df = result_df.join(repeat_counts, on="区域-受理号码", how="left")
        # 对地市进行简单排序，放在直接粘贴到重复投诉总表中
        result_df = result_df.sort("区域")

        filtered_repeat_df = result_df.filter(pl.col("重复投诉次数") >= 2)

        # 按照重复投诉次数降序排序
        filtered_repeat_df = filtered_repeat_df.sort("重复投诉次数")

        complaint_texts = []
        region_complaints = {}
        complaint_numbers = set()  # To track unique complaint numbers

        # 自定义区域排序
        region_order = ["南昌", "九江", "上饶", "抚州", "宜春", "吉安", "赣州", "景德镇", "萍乡", "新余", "鹰潭"]

        region_map = {region: idx for idx, region in enumerate(region_order)}

        region_complaints_data = {}  # 用来存储每个区域对应的投诉数据

        for row in filtered_repeat_df.iter_rows(named=True):
            complaint_number = row["区域-受理号码"]

            if complaint_number in complaint_numbers:
                continue

            complaint_numbers.add(complaint_number)

            complaint_count = row["重复投诉次数"]
            complaint_group = result_df.filter(pl.col("区域-受理号码") == complaint_number).sort("系统接单时间")

            text_content = [f"{complaint_number} ({complaint_count}次)"]

            for complaint_row in complaint_group.iter_rows(named=True):
                complaint_date = complaint_row["系统接单时间"]
                complaint_date_str = complaint_date.strftime("%m月%d日") if complaint_date else "日期无效"
                text_content.append(f"{complaint_date_str}投诉：用户来电反映在；")

            complaint_texts.append("\n".join(text_content))

            # 收集每个区域的投诉数据
            region = row["区域"].replace("市", "")  # Remove "市" from the region name
            if region not in region_complaints_data:
                region_complaints_data[region] = []
            region_complaints_data[region].append(
                {"complaint_number": complaint_number, "count": complaint_count, "text": "\n".join(text_content)})

            if region in region_map:
                region_complaints[region] = region_complaints.get(region, 0) + 1

        # 按照自定义顺序进行排序
        sorted_regions = sorted(region_complaints.items(), key=lambda x: region_map.get(x[0], float('inf')))

        region_summary = "、".join([f"{region}{count}单" for region, count in sorted_regions])

        # 排序后的文本内容
        complaint_texts_sorted = []
        for region, complaints in sorted_regions:
            # 每个区域内部按照重复投诉次数降序排序
            sorted_complaints = sorted(region_complaints_data[region], key=lambda x: x["count"], reverse=True)
            for complaint in sorted_complaints:
                complaint_texts_sorted.append(complaint["text"])

        # Concatenate all complaint texts into a single string with region summary at the top
        all_complaints_text = f"总共投诉：{region_summary}\n\n重复投诉内容如下:\n" + "\n".join(complaint_texts_sorted)

        # Create a new DataFrame for complaint text
        text_df = pl.DataFrame({"投诉信息": [all_complaints_text]})

        stats_df = generate_repeat_complaints_table(result_df)

        return dataframe, result_df, text_df, stats_df

    except Exception as e:
        logging.error(f"数据处理失败: {str(e)}")
        # 返回空的结果而不是直接raise
        empty_stats = pl.DataFrame({
            "区域": ["总计"],
            "重复2次": [0],
            "重复3次": [0],
            "重复4次及以上": [0],
            "当日新增重复投诉总计": [0],
            "今天重复投诉解决情况": [0],
            "累计重复投诉解决率": [None]
        })
        return df, pl.DataFrame(), pl.DataFrame({"投诉信息": ["处理失败"]}), empty_stats


if __name__ == '__main__':
    start_time = dt.datetime.now()
    file_manager = FileManager("WorkDocument\\重复投诉日报")

    source_file = file_manager.get_latest_file("source")

    # 读取Excel文件
    try:
        df = file_manager.read_excel(source_file)
    except Exception as e:
        logging.error(f"无法读取Excel文件: {e}")
        exit(1)

    # 处理数据并获取结果
    processed_df, result_df, text_df, stats_df = process_excel(df)

    file_manager.save_to_sheet("重复投诉日报", formatter=ExcelFormatter,  # 传入格式化器类
                               原始数据=processed_df, 重复投诉结果=result_df,
                               重复投诉文本=text_df, 重复投诉统计=stats_df)

    # 如果需要获取输出路径
    print(f"输出文件路径: {file_manager.output_path}")

    end_time = dt.datetime.now()
    runtime = end_time - start_time
    logging.info(f'运行时间：{runtime}')
