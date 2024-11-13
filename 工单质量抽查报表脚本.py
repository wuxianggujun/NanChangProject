import pandas as pd
import datetime as dt
import logging
import os
import Utils
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from pydantic import BaseModel, Field, model_validator, ValidationError, ValidationInfo
from typing import Optional


def get_digit_count(number):
    # 转换为整数，去掉小数部分
    return len(str(abs(int(number))))


class Coordinates(BaseModel):
    # 经度
    longitude: Optional[float] = Field(None, description="确认经度")
    # 纬度
    latitude: Optional[float] = Field(None, description="确认纬度")

    @staticmethod
    def validate_long_lat(longitude, latitude):
        # 检查是否为空或为0
        if (pd.isnull(longitude) or longitude == 0) and (pd.isnull(latitude) or latitude == 0):
            return "经纬度未填"
        if pd.isnull(longitude) or longitude == 0:
            return "经度未填"
        if pd.isnull(latitude) or latitude == 0:
            return "纬度未填"

            # 检查经纬度是否在对方的有效范围内
        if not (73.66 <= longitude <= 135.05 and 3.86 <= latitude <= 53.55):
            # print(f"经纬度填写为纬度： {longitude} : {latitude} ")
            # 判断是否经纬度填写错误：经度在纬度范围内，或者纬度在经度范围内
            if 3.86 <= longitude <= 53.55 and not (73.66 <= latitude <= 135.05):
                print(f"经度填写为纬度： {longitude} : {latitude} ")
                return "经度填写为纬度"
            if 73.66 <= latitude <= 135.05 and not (3.86 <= longitude <= 53.55):
                print(f"纬度填写为经度： {longitude} : {latitude} ")
                return "纬度填写为经度"
                # 判断是否经纬度互换：经度在纬度的有效范围内，纬度在经度的有效范围内
            if (3.86 <= longitude <= 53.55) and (73.66 <= latitude <= 135.05):
                print(f"经纬度互换： {longitude} : {latitude} ")
                return "经纬度互换"
            
        return None


def process_row(row):
    try:
        # 使用 Pydantic 模型进行验证
        error_message = Coordinates.validate_long_lat(longitude=row['确认经度'], latitude=row['确认纬度'])
        if error_message:
            return error_message
    except ValidationError as e:
        print("处理ROW错误：{e}")
    return None


def apply_error_flags(excel_data, index, error):
    """
    根据错误信息更新Excel数据中的相关列。
    """
    excel_data.at[index, '错误问题分类'] = '经纬度错误'
    excel_data.at[index, '错误问题描述'] = error
    excel_data.at[index, '是否抽查'] = '是'
    excel_data.at[index, '抽查是否有问题'] = '是'


def check_coordinate(row):
    longitude = row['确认经度']
    latitude = row['确认纬度']

    # 检查经纬度是否未填或为0
    if pd.isnull(latitude) or latitude == 0:
        return "纬度未填"
    elif pd.isnull(longitude) or longitude == 0:
        return "经度未填"

    # 检查是否存在经纬度互换的情况
    if -180 <= latitude <= 180 and -90 <= longitude <= 90:
        return "经纬度互换"

        # 检查小数点后位数（只在经纬度有效且不为空的情况下）
    if isinstance(longitude, float) and not pd.isnull(longitude):
        if len(str(longitude).split('.')[-1]) <= 4:
            return "经度小数点尾数小于等于4位"
    if isinstance(latitude, float) and not pd.isnull(latitude):
        if len(str(latitude).split('.')[-1]) <= 4:
            return "纬度小数点尾数小于等于4位"


def apply_header_style(output_path):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    wb = load_workbook(output_path)
    ws = wb.active

    # 设置新插入列的列头为黄色背景
    headers_to_color = ['回复客服内容', '是否反馈', '系统接单时间', '系统接单时间2', '反馈时长', '解决回复时间',
                        '是否现场处理', '用户是否在现场', '标签', '确认经度', '确认纬度', '问题类别', '行政场景',
                        '投诉场景', '问题区域', '是否解决', '定性类别', '处理类型', '用户认为是否及时响应',
                        '用户认为是否解决', '用户是否满意', '口碑未达情况原因', '错误问题分类', '错误问题描述',
                        '是否抽查', '抽查是否有问题']
    for cell in ws[1]:  # 第一行是列头
        if cell.value in headers_to_color:
            cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 水平和垂直居中
        # 设置第一行的行高为文字的两倍
    ws.row_dimensions[1].height = 30  # 调整此值以适应您的需求
    wb.save(output_path)
    wb.close()


def process_excel(df: pd.DataFrame):
    excel_data = df.copy()
    excel_data = excel_data.drop('序号', axis=1)
    # 通过客服流水号去除重复值
    excel_data = excel_data.drop_duplicates(subset=['客服流水号'])

    # 排除“处理类型中”包“退客服”的列
    excel_data = excel_data[~excel_data['处理类型'].str.contains('退客服', na=False)]

    # 在回复客服内容右侧新增一列是否反馈，默认为是
    excel_data.insert(excel_data.columns.get_loc('回复客服内容') + 1, '是否反馈', '是')
    # 第四步：在系统接单时间右侧新增系统接单时间2，设置格式只保留天数
    excel_data.insert(excel_data.columns.get_loc('系统接单时间') + 1, '系统接单时间2',
                      pd.to_datetime(excel_data['系统接单时间']).dt.strftime('%Y-%m-%d'))

    # 新增反馈时长列，格式为 hh:mm:ss
    excel_data.insert(
        excel_data.columns.get_loc('解决回复时间'), '反馈时长',
        (pd.to_datetime(excel_data['解决回复时间']) - pd.to_datetime(excel_data['系统接单时间']))
        .apply(lambda x: str(x).split('.')[0] if pd.notnull(x) else None)  # 去掉小数部分
    )

    print("读取到的列名：", excel_data.columns.tolist())

    # 第六步：在口碑未达情况原因右侧清除数据，并新增四列
    excel_data.insert(excel_data.columns.get_loc('口碑未达情况原因') + 1, '错误问题分类', None)
    excel_data.insert(excel_data.columns.get_loc('错误问题分类') + 1, '错误问题描述', None)
    excel_data.insert(excel_data.columns.get_loc('错误问题描述') + 1, '是否抽查', None)
    excel_data.insert(excel_data.columns.get_loc('是否抽查') + 1, '抽查是否有问题', None)

    # 第七步：删除“抽查是否有问题”右侧数据
    col_index = excel_data.columns.get_loc('抽查是否有问题')
    excel_data = excel_data.iloc[:, :col_index + 1]  # 保留“抽查是否有问题”列及其左侧所有列

    for index, row in excel_data.iterrows():

        # 检查这四类的经纬度是否正常
         if row['定性类别'] in ['负荷类', '优化类', '维护类', '建设类']:
            error_message = process_row(row)
            if error_message:
                apply_error_flags(excel_data, index, error_message)


    return excel_data 


if __name__ == '__main__':
    start_time = dt.datetime.now()
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    current_path = os.path.dirname(os.path.abspath(__file__))
    source_dir = os.path.join(current_path, "WorkDocument", "工单质量抽查报表", "source")

    if not os.path.exists(source_dir):
        logging.error("未找到指定目录，请检查路径是否正确！")
        exit(1)

    file_path = Utils.get_first_excel_file_in_dir(source_dir)

    logging.info(f"正在处理文件：{file_path}")

    df = Utils.read_excel(file_path)

    # 处理数据并获取结果
    processed_df = process_excel(df)

    # 保存处理后的数据到文件
    data_str = start_time.strftime("%Y%m%d")
    output_filename = f'工单质量抽查报表{data_str}{os.path.splitext(file_path)[1]}'
    output_path = os.path.join(current_path, "WorkDocument", "工单质量抽查报表", output_filename)

    # 保存原始数据
    Utils.save_to_excel(processed_df, output_path)

    apply_header_style(output_path)

    end_time = dt.datetime.now()
    runtime = end_time - start_time
    logging.info(f'运行时间：{runtime}')
