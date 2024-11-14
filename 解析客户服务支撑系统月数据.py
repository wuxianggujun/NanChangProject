import logging
import glob
import os.path

from tqdm import tqdm  # 导入tqdm
import polars as pl
from openpyxl.reader.excel import load_workbook

def process_excel(excel_data_df: pl.DataFrame):
    # 选择所需的列
    columns_to_keep = ['客服流水号', '受理号码', '区域', '系统接单时间', '月份']
    excel_data_df = excel_data_df.select(columns_to_keep)
    return excel_data_df


def standardize_column_types(df: pl.DataFrame) -> pl.DataFrame:
    # 在这里进行列类型统一的转换，假设我们希望所有列都转换为字符串类型
    for col in df.columns:
        # 假设将所有列类型统一为字符串，如果有其他需求，可以根据实际情况进行调整
        df = df.with_columns(pl.col(col).cast(pl.Utf8))
    return df


def process_dataframe(main_dataframe:pl.DataFrame):
    main_dataframe = main_dataframe.with_columns((pl.col("区域")+"-"+pl.col("受理号码")).alias("区域-受理号码"))
    
    for month in range(1, 11):
        main_dataframe = main_dataframe.with_columns(pl.lit(None).cast(pl.Utf8).alias(f"{month}月"))
    
    
    main_dataframe = main_dataframe.unique(subset=["客服流水号", "区域-受理号码", "月份"], keep="first")

    
    # 按“区域-受理号码”和“月份”分组，统计每个组合在该月的重复次数
    grouped = main_dataframe.group_by(["区域-受理号码", "月份"]).agg([
        pl.len().alias("重复次数")
    ])

    # 提取月份的数字部分，假设月份格式为 "YYYYMM"
    main_dataframe = main_dataframe.with_columns(
        pl.col("月份").str.slice(4, 2).cast(pl.Int32).alias("月份数字")
    )
    
    # 将“重复次数”添加回原数据框
    main_dataframe = main_dataframe.join(grouped, on=["区域-受理号码", "月份"], how="left")


    # main_dataframe = main_dataframe.unique(subset="客服流水号")
    main_dataframe = main_dataframe.unique(subset=["区域-受理号码", "月份"], keep="first")

    # 填充每一行对应月份的投诉次数
    for month in range(1, 11):
        # 检查当前行的“月份”，如果匹配，则填充“重复次数”到对应的月份列
        main_dataframe = main_dataframe.with_columns(
            pl.when(pl.col("月份数字").cast(pl.Utf8) == str(month))  # 如果月份匹配
            .then(pl.col("重复次数"))  # 填充重复次数
            .otherwise(pl.lit(None))  # 否则填充 None
            .alias(f"{month}月")  # 对应月份的列
        )

        # 按照“月份数字”排序，确保数据按月份顺序排列
    main_dataframe = main_dataframe.sort("月份数字")

    # 删除“重复次数”列，因为已经填充到月份列中了
    main_dataframe = main_dataframe.drop(["月份数字","重复次数"])
    
    return main_dataframe


def save_to_excel(dataframe: pl.DataFrame, file_path: str):
    # Convert Polars DataFrame to a list of dicts (rows)
    data = dataframe.to_dicts()

    if os.path.exists(file_path):
        os.remove(file_path)

    # Load the existing workbook or create a new one
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        workbook = Workbook()

    # Create a new sheet or use the existing one
    worksheet = workbook.active

    # Add column headers to the first row
    columns = dataframe.columns
    for col_num, col_name in enumerate(columns, start=1):
        worksheet.cell(row=1, column=col_num, value=col_name)

    # Add the data to the worksheet
    for row_num, row in tqdm(enumerate(data, start=2), total=len(data),desc="保存数据中"):
        for col_num, col_name in enumerate(columns, start=1):
            worksheet.cell(row=row_num, column=col_num, value=row[col_name])

    # Save the workbook
    workbook.save(file_path)
    print(f"File saved to {file_path}")

if __name__ == '__main__':
    try:
        folder_path = "WorkDocument/202401-10月支撑系统"
        file_list = glob.glob(folder_path + "/*.xls*")
    
        main_dataframe = pl.DataFrame()
    
        for file in tqdm(file_list, desc="读取文件列表中"):
            data_df = pl.read_excel(file)
            # 确保列类型一致
            data_df = standardize_column_types(data_df)
            data_df = process_excel(data_df)
            main_dataframe = pl.concat([main_dataframe, data_df],how="vertical")
            
        main_dataframe = process_dataframe(main_dataframe)
            
        save_to_excel(main_dataframe,"WorkDocument/全月份投诉明细.xlsx")
        
    except Exception as e:
        logging.error(e)
