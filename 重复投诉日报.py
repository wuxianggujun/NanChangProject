import datetime as dt
import logging
import polars as pl
from tool.data import DataUtils
from tool.file import FileManager
from tool.formatter import ExcelFormatter
import os
from openai import OpenAI  # 用于调用DeepSeek API
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side


def generate_repeat_complaints_table(dataframe: pl.DataFrame) -> pl.DataFrame:
    stats_data = []
    region_order = ["南昌", "九江", "上饶", "抚州", "宜春", "吉安", "赣州", "景德镇", "萍乡", "新余", "鹰潭"]

    # 初始化统计字典，为每个地区创建一个计数字典
    region_stats = {region: {"重复2次": 0, "重复3次": 0, "重复4次及以上": 0} for region in region_order}

    repeat_complaints = dataframe.filter(pl.col("重复投诉次数") >= 2)

    # 对每个区域-受理号码只统计一次
    unique_complaints = repeat_complaints.unique(subset=["区域-受理号码"])

    # 统计每个地区的重复投诉次数
    for row in unique_complaints.iter_rows(named=True):
        region = row["区域"].replace("市", "")
        count = row["重复投诉次数"]
       
        if region in region_stats:
            if count == 2:
                region_stats[region]["重复2次"] += 1
            elif count == 3:
                region_stats[region]["重复3次"] += 1
            elif count >= 4:
                region_stats[region]["重复4次及以上"] += 1
            
    # 转换为DataFrame格式
    for region in region_order:
        # 计算当日新增重复投诉总计
        total = (region_stats[region]["重复2次"] +
                 region_stats[region]["重复3次"] +
                 region_stats[region]["重复4次及以上"])

        # 将0值转换为None（这样在DataFrame中显示为空）
        row_data = {
            "区域": region,
            "重复2次": region_stats[region]["重复2次"] if region_stats[region]["重复2次"] > 0 else None,
            "重复3次": region_stats[region]["重复3次"] if region_stats[region]["重复3次"] > 0 else None,
            "重复4次及以上": region_stats[region]["重复4次及以上"] if region_stats[region][
                                                                          "重复4次及以上"] > 0 else None,
            "当日新增重复投诉总计": total if total > 0 else 0,
            "今天重复投诉解决情况": None,
            "累计重复投诉解决率": None
        }
        
        stats_data.append(row_data)

    # 添加总计行
    totals = {
        "区域": "总计",
        "重复2次": sum(row["重复2次"] or 0 for row in stats_data),
        "重复3次": sum(row["重复3次"] or 0 for row in stats_data),
        "重复4次及以上": sum(row["重复4次及以上"] or 0 for row in stats_data),
        "当日新增重复投诉总计": sum(row["当日新增重复投诉总计"] for row in stats_data),
        "今天重复投诉解决情况": 0,
        "累计重复投诉解决率": None
    }
    stats_data.append(totals)

    stats_df = pl.DataFrame(stats_data)
    return stats_df


def extract_address_and_issue(complaint_content):
    return "", ""


def extract_complaint_content(client, complaint_text):
    """
    使用DeepSeek API提取投诉内容的关键信息
    """
    try:
        prompt = f"""
        请从以下投诉记录中提取用户反映的具体问题，用简洁的语言描述：
        {complaint_text}
        只需要返回提取后的内容，不需要任何额外解释。
        """
        
        response = client.chat.completions.create(
            model="deepseek-chat",  # 使用适当的DeepSeek模型
            messages=[
                {"role": "system", "content": "你是一个专业的客服投诉内容提取助手，请简洁准确地提取用户投诉的具体问题。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=200
        )
        
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"AI提取文本失败: {str(e)}")
        return "用户反映在"  # 修改了默认返回值


def process_excel(df: pl.DataFrame) -> tuple:
    # 初始化DeepSeek API客户端
    try:
        client = OpenAI(
            api_key=os.getenv("DEEPSEEK_API_KEY"),
            base_url="https://api.deepseek.com/v1"  # DeepSeek的API地址
        )
    except Exception as e:
        logging.error(f"初始化DeepSeek API客户端失败: {str(e)}")
        client = None

    # 打印初始数据信息
    logging.info(f"初始数据形状: {df.shape}")
    logging.info(f"系统接单时间列类型: {df['系统接单时间'].dtype}")
    logging.info(f"系统接单时间样本: {df['系统接单时间'].head()}")

    try:
      
        logging.info("初始数据类型转换完成")

            # 定义时间范围
        end_time = dt.datetime.now().replace(hour=16, minute=0, second=0, microsecond=0)
        start_time = (end_time - dt.timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
      
        logging.info(f"开始处理：{start_time} 到 {end_time} 共三十天的数据...")

        dataframe = DataUtils(df).filter_data_range(
            date_column="系统接单时间",
            start_time=start_time,
            end_time=end_time
        )

        dataframe = DataUtils(dataframe).clean_and_unique(
            unique_columns=["客服流水号"],
            drop_columns=["序号", "工单号"]
        )

        dataframe = DataUtils(dataframe).add_date_only_column(
            date_column="系统接单时间",
            new_column="系统接单时间2",
            format="%Y/%m/%d %H"
        )

        dataframe = DataUtils(dataframe).insert_colum("系统接单时间", "系统接单时间2")

        dataframe = DataUtils(dataframe).drop_columns_after(column_name="口碑未达情况原因")

        dataframe = DataUtils(dataframe).combine_columns(
            columns=["区域", "受理号码"],
            separator="-",
            new_column="区域-受理号码"
        )

        # 添加筛选条件：今天下午四点到昨天下午四点的数据
        yesterday_end = (dt.datetime.now() - dt.timedelta(days=1)).replace(hour=16, minute=0, second=0, microsecond=0)
        today_start = dt.datetime.now().replace(hour=16, minute=0, second=0, microsecond=0)

        # 筛选出时间段内的重复工单
        filtered_df = dataframe.filter(
            (pl.col("系统接单时间") >= yesterday_end) & (pl.col("系统接单时间") <= today_start)
        )

        # 检查是否有数据
        if len(filtered_df) == 0:
            logging.warning(f"在 {yesterday_end} 到 {today_start} 期间没有找到任何投诉数据")
            # 返回空的DataFrame但保持结构一致
            empty_stats = pl.DataFrame({
                "区域": ["总计"],
                "重复2次": [0],
                "重复3次": [0],
                "重复4次及以上": [0],
                "当日新增重复投诉总计": [0],
                "今天重复投诉解决情况": [0],
                "累计重复投诉解决率": [None]
            })
            return dataframe, pl.DataFrame(), pl.DataFrame({"投诉信息": ["无重复投诉数据"]}), empty_stats, None, None

        result_list = []

        for row in filtered_df.iter_rows(named=True):
            matches = dataframe.filter(pl.col("区域-受理号码") == row["区域-受理号码"])
            result_list.extend(matches.to_dicts())

        result_df = pl.DataFrame(result_list)

        result_df = DataUtils(result_df).clean_and_unique(
            unique_columns=["客服流水号"]
        )
        
         # 计算重复投诉次数
        repeat_counts = result_df.group_by("区域-受理号码").agg(pl.len().alias("重复投诉次数"))

        result_df = result_df.join(repeat_counts, on="区域-受理号码", how="left")
        # 对地市进行简单排序，放在直接粘贴到重复投诉总表中
        result_df = result_df.sort("区域")

        filtered_repeat_df = result_df.filter(pl.col("重复投诉次数") >= 2)

        # 按照重复投诉次数降序排序
        filtered_repeat_df = filtered_repeat_df.sort("重复投诉次数")

        complaint_texts = []
        region_complaints = {}
        complaint_numbers = set()  # To track unique complaint numbers

        # 自定义区域排序
        region_order = ["南昌", "九江", "上饶", "抚州", "宜春", "吉安", "赣州", "景德镇", "萍乡", "新余", "鹰潭"]

        region_map = {region: idx for idx, region in enumerate(region_order)}

        region_complaints_data = {}  # 用来存储每个区域对应的投诉数据

        for row in filtered_repeat_df.iter_rows(named=True):
            complaint_number = row["区域-受理号码"]

            if complaint_number in complaint_numbers:
                continue

            complaint_numbers.add(complaint_number)

            complaint_count = row["重复投诉次数"]
            complaint_group = result_df.filter(pl.col("区域-受理号码") == complaint_number).sort("系统接单时间")

            text_content = [f"{complaint_number} ({complaint_count}次)"]

            for complaint_row in complaint_group.iter_rows(named=True):
                complaint_date = complaint_row["系统接单时间"]
                complaint_date_str = complaint_date.strftime("%m月%d日") if complaint_date else "日期无效"
                service_number = complaint_row["客服流水号"]
                
                # 获取原始投诉内容
                original_complaint = complaint_row.get("投诉内容", "")  # 假设有"投诉内容"列
                
                # 使用AI提取投诉内容
                if client and original_complaint:
                    extracted_content = extract_complaint_content(client, original_complaint)
                else:
                    extracted_content = "用户反映在"
                    
                text_content.append(f"[{service_number}]{complaint_date_str}投诉：{extracted_content}；")

            complaint_texts.append("\n".join(text_content))

            # 收集每个区域的投诉数据
            region = row["区域"].replace("市", "")  # Remove "市" from the region name
            if region not in region_complaints_data:
                region_complaints_data[region] = []
            region_complaints_data[region].append(
                {"complaint_number": complaint_number, "count": complaint_count, "text": "\n".join(text_content)})

            if region in region_map:
                region_complaints[region] = region_complaints.get(region, 0) + 1

        # 按照自定义顺序进行排序
        sorted_regions = sorted(region_complaints.items(), key=lambda x: region_map.get(x[0], float('inf')))

        # 获取当天日期并格式化
        today_date = dt.datetime.now().strftime("%Y%m%d")
        region_summary = f"{today_date}新增重复投诉：" + "、".join([f"{region}{count}单" for region, count in sorted_regions]) + "；"

        # 排序后的文本内容
        complaint_texts_sorted = []
        for region, complaints in sorted_regions:
            # 每个区域内部按照重复投诉次数降序排序
            sorted_complaints = sorted(region_complaints_data[region], key=lambda x: x["count"], reverse=True)
            for complaint in sorted_complaints:
                complaint_texts_sorted.append(complaint["text"])

        # Concatenate all complaint texts into a single string with region summary at the top
        all_complaints_text = f"{region_summary}\n\n重复投诉内容如下:\n" + "\n".join(complaint_texts_sorted)

        # Create a new DataFrame for complaint text
        text_df = pl.DataFrame({f"{dt.datetime.now().strftime('%Y%m%d')}新增重复投诉": [all_complaints_text]})

        stats_df = generate_repeat_complaints_table(result_df)

        # 在返回结果前，添加重复投诉总表的数据处理
        repeat_total_df = None
        repeat_sheet_df = None  # 用于保存到sheet的数据
        if len(filtered_repeat_df) > 0:
            # 获取当前日期相关信息
            today = dt.datetime.now()
            date_str = today.strftime("%Y%m%d")
            month_str = f"{today.month}月"
            
            # 准备重复投诉总表数据
            repeat_total_df = filtered_repeat_df.clone()  # 克隆原始数据
            repeat_sheet_df = filtered_repeat_df.clone()  # 克隆用于sheet的数据
            
            # 先添加新列
            repeat_total_df = repeat_total_df.with_columns([
                pl.col("区域-受理号码").alias("重复投诉号码"),
                pl.lit(date_str).alias("日期"),
                pl.lit(month_str).alias("月份")
            ])
            
            repeat_sheet_df = repeat_sheet_df.with_columns([
                pl.col("区域-受理号码").alias("重复投诉号码"),
                pl.lit(date_str).alias("日期"),
                pl.lit(month_str).alias("月份")
            ])
            
            # 重新排列列的顺序，确保三列在最后并按正确顺序排列
            original_cols = [col for col in repeat_total_df.columns if col not in ["重复投诉号码", "日期", "月份"]]
            new_cols = ["重复投诉号码", "日期", "月份"]
            
            repeat_total_df = repeat_total_df.select(original_cols + new_cols)
            repeat_sheet_df = repeat_sheet_df.select(original_cols + new_cols)

        # 返回所有处理后的数据，包括新增的sheet数据
        return dataframe, result_df, text_df, stats_df, repeat_total_df, repeat_sheet_df

    except Exception as e:
        logging.error(f"数据处理失败: {str(e)}")
        # 返回空的结果而不是直接raise
        empty_stats = pl.DataFrame({
            "区域": ["总计"],
            "重复2次": [0],
            "重复3次": [0],
            "重复4次及以上": [0],
            "当日新增重复投诉总计": [0],
            "今天重复投诉解决情况": [0],
            "累计重复投诉解决率": [None]
        })
        return df, pl.DataFrame(), pl.DataFrame({"投诉信息": ["处理失败"]}), empty_stats, None, None


def export_repeat_complaints_to_excel(repeat_total_df: pl.DataFrame, file_manager: FileManager):
    """
    导出重复投诉数据到指定目录的Excel文件中
    """
    try:
        if repeat_total_df is None or len(repeat_total_df) == 0:
            logging.info("没有重复投诉数据需要导出")
            return

        # 创建重复投诉总表目录
        total_complaints_dir = os.path.join(file_manager.base_dir, "重复投诉日报总表")
        os.makedirs(total_complaints_dir, exist_ok=True)
        
        # 文件路径
        file_path = os.path.join(total_complaints_dir, "重复投诉总表.xlsx")
        
        # 打开现有文件
        if not os.path.exists(file_path):
            logging.error(f"重复投诉总表文件不存在: {file_path}")
            return
            
        workbook = load_workbook(file_path)
        sheet = workbook.active if "重复投诉总表数据" not in workbook.sheetnames else workbook["重复投诉总表数据"]
        
        # 检查是否已有当天的数据
        date_str = repeat_total_df["日期"][0]
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_idx, column=2).value) == date_str:
                logging.info(f"已存在{date_str}的数据，不再添加")
                return
        
        # 获取最后一行的位置
        start_row = sheet.max_row + 1

        # 设置样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row in repeat_total_df.iter_rows(named=True):
            current_col = 1
            for col_name in repeat_total_df.columns:
                value = row[col_name]
                if isinstance(value, dt.datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                cell = sheet.cell(row=start_row, column=current_col, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment
                current_col += 1
            start_row += 1

        # 保存文件
        workbook.save(file_path)
        logging.info(f"重复投诉数据已追加到文件: {file_path}")
        
    except Exception as e:
        logging.error(f"导出重复投诉数据失败: {str(e)}")


if __name__ == '__main__':
    start_time = dt.datetime.now()
    file_manager = FileManager("WorkDocument\\重复投诉日报")

    source_file = file_manager.get_latest_file("source")

    # 读取Excel文件
    try:
        df = file_manager.read_excel(file_path=source_file)
    except Exception as e:
        logging.error(f"无法读取Excel文件: {e}")
        exit(1)

    # 处理数据并获取结果
    processed_df, result_df, text_df, stats_df, repeat_total_df, repeat_sheet_df = process_excel(df)

    # 保存原有的sheet，并添加新的重复投诉总表sheet
    file_manager.save_to_sheet("重复投诉日报", formatter=ExcelFormatter,
                              原始数据=processed_df, 
                              重复投诉结果=result_df,
                              重复投诉文本=text_df, 
                              重复投诉统计=stats_df,
                              重复投诉日报总表=repeat_sheet_df)  # 添加新的sheet

    # 导出重复投诉数据到单独的文件
    export_repeat_complaints_to_excel(repeat_total_df, file_manager)

    end_time = dt.datetime.now()
    runtime = end_time - start_time
    logging.info(f'运行时间：{runtime}')
