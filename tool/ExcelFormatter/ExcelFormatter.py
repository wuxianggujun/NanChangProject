from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.default_font = Font(name='宋体', size=10)
        self.header_font = Font(name='宋体',size=12,bold=True)
        self.bold_font = Font(name='宋体',size=12,bold=True)

    def format_repeat_complaints_table(self,sheet_name: str):
        """
            格式化重复投诉统计表
        """
        ws = self.workbook[sheet_name]

        column_widths = {}

        # # 设置列宽
        # column_widths = {
        #     'A': 12,  # 区域
        #     'B': 12,  # 重复2次
        #     'C': 12,  # 重复3次
        #     'D': 16,  # 重复4次及以上
        #     'E': 20,  # 当日新增重复投诉总计
        #     'F': 20,  # 今天重复投诉解决情况
        #     'G': 20,  # 累计重复投诉解决率
        # }


        # 用于跟踪每列的最大宽度
        column_widths = {}
    
        # 遍历所有单元格计算所需的列宽
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # 获取列号
                    col = cell.column_letter
                    # 计算当前值的显示宽度（中文字符计为2个单位，其他字符计为1个单位）
                    str_value = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in str_value)
                
                    # 更新该列的最大宽度
                    current_width = column_widths.get(col, 0)
                    column_widths[col] = max(current_width, width)
          
        # 设置列宽（添加一些padding）
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = max(width+4,10)
            
        # 设置统一行高（单位：磅）
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20  # 可以调整这个值来改变行高


        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        alignment = Alignment(horizontal='center', vertical='center')
      

        for cell in ws[1]:
            cell.border = border
            cell.alignment = alignment
            cell.font = self.header_font

        data_rows = list(ws.rows)
        max_row = len(data_rows)
        max_col = len(data_rows[0])

        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.default_font
                cell.border = border
                cell.alignment = alignment

                 
                # 设置当日新增重复投诉总计列（E列）大于0的数字为粗体
                if col == 5:  # E列
                    try:
                        value = float(cell.value) if cell.value else 0
                        if value > 0:
                            cell.font = self.bold_font
                    except (ValueError, TypeError):
                        pass

                if col == 7 and cell.value:
                    try:
                        value = float(cell.value.strip('%')) / 100
                        cell.value = value
                        cell.number_format = '0.00%'
                    except (ValueError, AttributeError):
                        pass

        for cell in ws[max_row]:
            cell.font = self.header_font

    def apply_formatting(self, sheet_names: str = None):
        """
            应用格式化到指定的Sheet或所有Sheet
        """
        if sheet_names is None:
            sheet_names = self.workbook.sheetnames
        elif isinstance(sheet_names, str):
            sheet_names = [sheet_names]

        for sheet_name in sheet_names:
            if "重复投诉统计" in sheet_name:
                self.format_repeat_complaints_table(sheet_name)




